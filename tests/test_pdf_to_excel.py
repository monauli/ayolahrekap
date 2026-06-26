import sys
import tempfile
import types
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from app import convert_pdfs_to_excel


class _FakePage:
    def __init__(self, tables=None, text=""):
        self._tables = tables or []
        self._text = text

    def extract_tables(self):
        return self._tables

    def extract_text(self, layout=False):
        return self._text


class _FakePdf:
    def __init__(self, pages):
        self.pages = pages

    def __enter__(self):
        return self

    def __exit__(self, *_args):
        return False


class PdfToExcelTests(unittest.TestCase):
    def test_tables_and_text_are_written_to_separate_page_sheets(self):
        pages = [
            _FakePage(tables=[[['Produk', 'Jumlah'], ['Raket', '2'], ['=SUM(A1)', '1']]]),
            _FakePage(text="Catatan akhir\nTotal  300000"),
        ]
        fake_module = types.SimpleNamespace(open=lambda _path: _FakePdf(pages))

        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "laporan.pdf"
            source.touch()
            output = Path(temp_dir) / "hasil.xlsx"
            with patch.dict(sys.modules, {"pdfplumber": fake_module}):
                result_path, stats = convert_pdfs_to_excel([source], output_path=output)

            workbook = load_workbook(result_path, data_only=False)
            self.assertEqual(workbook.sheetnames, ["Ringkasan", "laporan 1", "laporan 2"])
            self.assertEqual(stats["page_count"], 2)
            self.assertEqual(stats["table_count"], 1)
            self.assertEqual(stats["row_count"], 5)
            self.assertEqual(workbook["laporan 1"]["A3"].value, "'=SUM(A1)")
            self.assertEqual(workbook["laporan 1"]["A1"].fill.fgColor.rgb, "000F4A5D")
            self.assertEqual(workbook["laporan 2"]["A2"].value, "Total")
            self.assertEqual(workbook["laporan 2"]["B2"].value, "300000")

    def test_image_only_pdf_returns_actionable_error(self):
        fake_module = types.SimpleNamespace(open=lambda _path: _FakePdf([_FakePage()]))
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "scan.pdf"
            source.touch()
            with patch.dict(sys.modules, {"pdfplumber": fake_module}):
                with self.assertRaisesRegex(ValueError, "OCR"):
                    convert_pdfs_to_excel([source], output_path=Path(temp_dir) / "hasil.xlsx")


if __name__ == "__main__":
    unittest.main()
