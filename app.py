from __future__ import annotations

import calendar
import hashlib
import json
import re
import sqlite3
import sys
from collections import defaultdict
from contextlib import contextmanager
from copy import copy
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any, Iterator

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


if getattr(sys, "frozen", False):
    BASE_DIR = Path(sys.executable).resolve().parent
else:
    BASE_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = BASE_DIR / "outputs"
DATA_DIR = BASE_DIR / "data"
OUTPUT_DIR.mkdir(exist_ok=True)
DATA_DIR.mkdir(exist_ok=True)
DB_PATH = DATA_DIR / "rekap.db"

RAW_HEADERS = [
    "No",
    "Booking ID",
    "Venue",
    "Court",
    "Court ID",
    "Sports Category",
    "Customer Name",
    "Customer Email",
    "Customer Phone",
    "Username Customer",
    "Date of Booking",
    "Booking Period Start Time",
    "Booking Period End Time",
    "Session Length",
    "Price",
    "Payment ID",
    "Payment Method",
    "Total Booking Amount",
    "Discount",
    "Ayo Discount",
    "Voucher",
    "Net Booking Amount",
    "First/Down Payment",
    "Pelunasan Online",
    "Pelunasan Offline",
    "Status",
    "Revenue Venue",
    "Date Revenue Processed",
    "Note",
    "Booking Oleh",
    "Tanggal dan Waktu Booking",
    "Reschedule Oleh",
    "Tanggal dan Waktu Reschedule",
    "Dibatalkan Oleh",
    "Tanggal dan Waktu Pembatalan",
    "Alasan Pembatalan",
]

# Dikosongkan: dulu berisi 2 Booking ID MN spesifik dari workbook acuan dev yang ikut
# ke-exclude di SEMUA file. Sekarang exclusion hanya dari status (Failed/Cancelled).
DEFAULT_EXCLUDED_BOOKING_IDS: dict[str, str] = {}

# Reason yang dipakai saat dev nge-seed exclusion bawaan (dibersihkan otomatis di init_db).
LEGACY_SEEDED_EXCLUSION_REASON = "Payment Fail pada workbook acuan"

MONTH_NAMES = {
    1: "Januari",
    2: "Februari",
    3: "Maret",
    4: "April",
    5: "Mei",
    6: "Juni",
    7: "Juli",
    8: "Agustus",
    9: "September",
    10: "Oktober",
    11: "November",
    12: "Desember",
}

MONTH_ABBREV = {
    1: "Jan",
    2: "Feb",
    3: "Mar",
    4: "Apr",
    5: "Mei",
    6: "Jun",
    7: "Jul",
    8: "Agu",
    9: "Sep",
    10: "Okt",
    11: "Nov",
    12: "Des",
}

MONEY_HEADERS = {
    "Price",
    "Total Booking Amount",
    "Discount",
    "Ayo Discount",
    "Net Booking Amount",
    "First/Down Payment",
    "Pelunasan Online",
    "Pelunasan Offline",
    "Revenue Venue",
}

IMPORTANT_HEADERS = {
    "Booking ID",
    "Court",
    "Date of Booking",
    "Booking Period Start Time",
    "Booking Period End Time",
    "Payment Method",
    "Revenue Venue",
    "Status",
}


class ProcessCancelled(Exception):
    """Raised when the desktop app asks a long-running rekap to stop."""


def ensure_not_cancelled(should_cancel=None) -> None:
    if should_cancel and should_cancel():
        raise ProcessCancelled("Proses dibatalkan.")


def normalize_header(value: Any) -> str:
    text = "" if value is None else str(value)
    return re.sub(r"\s+", " ", text).strip()


def as_key(value: Any) -> str:
    return normalize_header(value).lower()


# Pra-cek cepat: hindari strptime untuk teks yang jelas bukan tanggal/jam (besar pengaruhnya
# saat membaca file ribuan baris — nama, Booking ID, dll tidak perlu dicoba 10+ format).
_DATE_HINT_RE = re.compile(r"^\s*(\d{1,4}[-/]\d{1,2}[-/]\d{1,4}|[A-Za-z]{3,9}\.?\s+\d{1,2},\s*\d{4})")
_TIME_HINT_RE = re.compile(r"^\s*\d{1,2}:\d{2}(:\d{2})?\s*$")


def parse_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()
        except OverflowError:
            return None

    text = normalize_header(value)
    if not _DATE_HINT_RE.match(text):
        return None
    for fmt in (
        "%Y-%m-%d",
        "%Y-%m-%d %H:%M:%S",
        "%d/%m/%Y",
        "%d/%m/%Y %H:%M:%S",
        "%m/%d/%Y",
        "%m/%d/%Y %H:%M:%S",
        "%d-%m-%Y",
        "%d-%m-%Y %H:%M:%S",
        "%b %d, %Y",
        "%B %d, %Y",
    ):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def parse_time_value(value: Any) -> time | None:
    if isinstance(value, datetime):
        return value.time().replace(microsecond=0)
    if isinstance(value, time):
        return value.replace(microsecond=0)
    if isinstance(value, (int, float)) and 0 <= value < 1:
        seconds = int(round(float(value) * 24 * 60 * 60))
        return (datetime.min + timedelta(seconds=seconds)).time()

    text = normalize_header(value)
    if not _TIME_HINT_RE.match(text):
        return None
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(text, fmt).time()
        except ValueError:
            pass
    return None


def format_excel_value(value: Any) -> Any:
    parsed_date = parse_date(value)
    if parsed_date and not isinstance(value, (int, float)):
        original = normalize_header(value)
        if re.fullmatch(r"\d{4}-\d{2}-\d{2}", original):
            return original

    parsed_time = parse_time_value(value)
    if parsed_time:
        return parsed_time.strftime("%H:%M:%S")

    if isinstance(value, str):
        return normalize_header(value)
    return value


def number_value(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = normalize_header(value).replace(",", "")
    try:
        return float(text)
    except ValueError:
        return 0.0


@contextmanager
def db_connect() -> Iterator[sqlite3.Connection]:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        with conn:
            yield conn
    finally:
        conn.close()


BOOKINGS_SCHEMA = """
CREATE TABLE IF NOT EXISTS bookings (
    row_key TEXT PRIMARY KEY,
    booking_id TEXT NOT NULL COLLATE NOCASE,
    booking_date TEXT,
    booking_year INTEGER,
    booking_month INTEGER,
    booking_day INTEGER,
    court TEXT,
    start_time TEXT,
    payment_method TEXT,
    channel TEXT,
    revenue REAL,
    fields_json TEXT NOT NULL,
    source_filename TEXT,
    uploaded_at TEXT NOT NULL
)
"""

BOOKING_ROW_IDENTITY_HEADERS = (
    "No",
    "Booking ID",
    "Venue",
    "Court",
    "Court ID",
    "Sports Category",
    "Customer Name",
    "Customer Phone",
    "Date of Booking",
    "Booking Period Start Time",
    "Booking Period End Time",
    "Session Length",
    "Price",
    "Payment ID",
    "Payment Method",
    "Total Booking Amount",
    "Discount",
    "Ayo Discount",
    "Voucher",
    "Net Booking Amount",
    "First/Down Payment",
    "Pelunasan Online",
    "Pelunasan Offline",
    "Status",
    "Revenue Venue",
    "Date Revenue Processed",
    "Booking Oleh",
    "Tanggal dan Waktu Booking",
    "Reschedule Oleh",
    "Tanggal dan Waktu Reschedule",
)


def row_key_value(header: str, value: Any) -> str:
    if header == "Booking ID":
        return normalize_header(value).upper()
    if header == "Date of Booking":
        parsed = parse_date(value)
        return parsed.isoformat() if parsed else normalize_header(value)
    if header in {"Booking Period Start Time", "Booking Period End Time"}:
        parsed_time = parse_time_value(value)
        return parsed_time.strftime("%H:%M:%S") if parsed_time else normalize_header(value)
    if header in MONEY_HEADERS:
        return f"{number_value(value):.2f}"
    return normalize_header(value)


def record_row_key(record: dict[str, Any], fallback: str | None = None) -> str:
    """Kunci stabil untuk satu baris export booking.

    Booking ID saja tidak cukup karena export AYO kadang berisi Booking ID yang
    sama di beberapa baris dengan nominal berbeda. Semua baris sumber harus ikut
    rekap, jadi identitas memakai detail baris dan posisi baris sumber.
    """
    identity = {
        header: row_key_value(header, record.get(header))
        for header in BOOKING_ROW_IDENTITY_HEADERS
    }
    identity["_source_row"] = normalize_header(record.get("_source_row"))
    if fallback is not None:
        identity["_fallback"] = normalize_header(fallback)
    encoded = json.dumps(identity, ensure_ascii=False, sort_keys=True)
    return hashlib.sha1(encoded.encode("utf-8")).hexdigest()


def booking_table_has_unique_booking_id(conn: sqlite3.Connection) -> bool:
    table = conn.execute(
        "SELECT sql FROM sqlite_master WHERE type = 'table' AND name = 'bookings'"
    ).fetchone()
    table_sql = str(table["sql"] or "") if table else ""
    if re.search(r"booking_id\s+TEXT\s+NOT\s+NULL\s+COLLATE\s+NOCASE\s+UNIQUE", table_sql, re.I):
        return True

    for index in conn.execute("PRAGMA index_list(bookings)").fetchall():
        if not int(index["unique"]):
            continue
        index_name = index["name"]
        indexed_columns = [
            normalize_header(row["name"]).lower()
            for row in conn.execute(f"PRAGMA index_info({index_name})").fetchall()
        ]
        if indexed_columns == ["booking_id"]:
            return True
    return False


def migrate_bookings_table_if_needed(conn: sqlite3.Connection) -> None:
    exists = conn.execute(
        "SELECT 1 FROM sqlite_master WHERE type = 'table' AND name = 'bookings'"
    ).fetchone()
    if not exists:
        conn.execute(BOOKINGS_SCHEMA)
        return

    columns = [row["name"] for row in conn.execute("PRAGMA table_info(bookings)").fetchall()]
    expected_columns = {
        "row_key",
        "booking_id",
        "booking_date",
        "booking_year",
        "booking_month",
        "booking_day",
        "court",
        "start_time",
        "payment_method",
        "channel",
        "revenue",
        "fields_json",
        "source_filename",
        "uploaded_at",
    }
    if expected_columns <= set(columns) and not booking_table_has_unique_booking_id(conn):
        return

    conn.execute("ALTER TABLE bookings RENAME TO bookings_old")
    conn.execute(BOOKINGS_SCHEMA)
    used_row_keys: set[str] = set()
    old_rows = conn.execute(
        "SELECT rowid AS old_rowid, * FROM bookings_old ORDER BY uploaded_at, rowid"
    ).fetchall()
    for row in old_rows:
        fields = json.loads(row["fields_json"])
        fields.setdefault("Booking ID", row["booking_id"])
        row_key = record_row_key(fields)
        fallback_counter = 0
        while row_key in used_row_keys:
            fallback_counter += 1
            row_key = record_row_key(fields, f"migration:{row['old_rowid']}:{fallback_counter}")
        used_row_keys.add(row_key)
        conn.execute(
            """
            INSERT INTO bookings (
                row_key, booking_id, booking_date, booking_year, booking_month, booking_day,
                court, start_time, payment_method, channel, revenue,
                fields_json, source_filename, uploaded_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                row_key,
                row["booking_id"],
                row["booking_date"],
                row["booking_year"],
                row["booking_month"],
                row["booking_day"],
                row["court"],
                row["start_time"],
                row["payment_method"],
                row["channel"],
                row["revenue"],
                row["fields_json"],
                row["source_filename"],
                row["uploaded_at"],
            ),
        )
    conn.execute("DROP TABLE bookings_old")


def init_db() -> None:
    with db_connect() as conn:
        migrate_bookings_table_if_needed(conn)
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS exclusions (
                booking_id TEXT PRIMARY KEY,
                reason TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS process_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                processed_at TEXT NOT NULL,
                report_date TEXT,
                report_month TEXT NOT NULL,
                output_file TEXT NOT NULL,
                raw_rows INTEGER NOT NULL,
                included_rows INTEGER NOT NULL,
                excluded_rows INTEGER NOT NULL,
                ayo_rows INTEGER NOT NULL,
                walk_in_rows INTEGER NOT NULL,
                total_revenue REAL NOT NULL,
                ayo_revenue REAL NOT NULL,
                walk_in_revenue REAL NOT NULL
            )
            """
        )
        log_columns = [row["name"] for row in conn.execute("PRAGMA table_info(process_logs)").fetchall()]
        for column_name, column_type in {
            "source_file": "TEXT",
            "feature": "TEXT",
            "period_label": "TEXT",
        }.items():
            if column_name not in log_columns:
                conn.execute(f"ALTER TABLE process_logs ADD COLUMN {column_name} {column_type}")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_bookings_date ON bookings (booking_date)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_bookings_month ON bookings (booking_year, booking_month)")
        conn.execute("DROP INDEX IF EXISTS idx_bookings_booking_id")
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_bookings_booking_id "
            "ON bookings (booking_id COLLATE NOCASE)"
        )
        conn.execute("CREATE INDEX IF NOT EXISTS idx_process_logs_date ON process_logs (report_date)")
        # Bersihkan exclusion bawaan dev yang dulu sempat di-seed, supaya database lama
        # ikut sembuh otomatis (mis. transaksi 'daniel' MN/.../0001887 & 0001888).
        conn.execute("DELETE FROM exclusions WHERE reason = ?", (LEGACY_SEEDED_EXCLUSION_REASON,))
        for booking_id, reason in DEFAULT_EXCLUDED_BOOKING_IDS.items():
            conn.execute(
                "INSERT OR IGNORE INTO exclusions (booking_id, reason) VALUES (?, ?)",
                (booking_id, reason),
            )


def save_records_to_db(records: list[dict[str, Any]], source_filename: str, should_cancel=None) -> int:
    init_db()
    uploaded_at = datetime.now().isoformat(timespec="seconds")
    saved = 0
    prepared_records: list[tuple[int, dict[str, Any], str, date]] = []
    upload_dates: set[date] = set()
    for index, record in enumerate(records):
        if index % 100 == 0:
            ensure_not_cancelled(should_cancel)
        booking_id = normalize_header(record.get("Booking ID"))
        booking_date = parse_date(record.get("Date of Booking"))
        if not booking_id or not booking_date:
            continue
        prepared_records.append((index, record, booking_id, booking_date))
        upload_dates.add(booking_date)

    with db_connect() as conn:
        if upload_dates:
            placeholders = ", ".join("?" for _ in upload_dates)
            conn.execute(
                f"DELETE FROM bookings WHERE booking_date IN ({placeholders})",
                [upload_date.isoformat() for upload_date in sorted(upload_dates)],
            )

        for index, record, booking_id, booking_date in prepared_records:
            if index % 100 == 0:
                ensure_not_cancelled(should_cancel)

            start_time = parse_time_value(record.get("Booking Period Start Time"))
            channel = booking_channel(record)
            payload = {header: record.get(header) for header in RAW_HEADERS if header in record}
            payload["_source_row"] = record.get("_source_row")
            row_key = record_row_key(record, f"upload:{index}")
            conn.execute(
                """
                INSERT OR REPLACE INTO bookings (
                    row_key, booking_id, booking_date, booking_year, booking_month, booking_day,
                    court, start_time, payment_method, channel, revenue,
                    fields_json, source_filename, uploaded_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    row_key,
                    booking_id,
                    booking_date.isoformat(),
                    booking_date.year,
                    booking_date.month,
                    booking_date.day,
                    normalize_header(record.get("Court")),
                    start_time.strftime("%H:%M:%S") if start_time else "",
                    normalize_header(record.get("Payment Method")),
                    channel,
                    number_value(record.get("Revenue Venue")),
                    json.dumps(payload, ensure_ascii=False, default=str),
                    source_filename,
                    uploaded_at,
                ),
            )
            saved += 1
        ensure_not_cancelled(should_cancel)
    return saved


def load_records_from_db(
    report_date: date | None,
    report_month: date | None,
    report_dates: list[date] | None = None,
) -> list[dict[str, Any]]:
    init_db()
    params: list[Any] = []
    where = ""
    if report_dates:
        placeholders = ", ".join("?" for _ in report_dates)
        where = f"WHERE b.booking_date IN ({placeholders})"
        params.extend(d.isoformat() for d in report_dates)
    elif report_date:
        where = "WHERE b.booking_date = ?"
        params.append(report_date.isoformat())
    elif report_month:
        where = "WHERE b.booking_year = ? AND b.booking_month = ?"
        params.extend([report_month.year, report_month.month])

    query = f"""
        SELECT b.fields_json, e.reason AS exclusion_reason
        FROM bookings b
        LEFT JOIN exclusions e ON e.booking_id = b.booking_id
        {where}
        ORDER BY b.booking_date, b.court, b.start_time, b.booking_id
    """
    with db_connect() as conn:
        rows = conn.execute(query, params).fetchall()

    records: list[dict[str, Any]] = []
    for row in rows:
        record = json.loads(row["fields_json"])
        if row["exclusion_reason"]:
            record["_excluded_reason"] = row["exclusion_reason"]
        records.append(record)
    return records


def db_record_count() -> int:
    init_db()
    with db_connect() as conn:
        return int(conn.execute("SELECT COUNT(*) FROM bookings").fetchone()[0])


def clear_process_logs() -> int:
    init_db()
    with db_connect() as conn:
        count = int(conn.execute("SELECT COUNT(*) FROM process_logs").fetchone()[0])
        conn.execute("DELETE FROM process_logs")
    return count


def reset_database() -> dict[str, int]:
    init_db()
    with db_connect() as conn:
        booking_count = int(conn.execute("SELECT COUNT(*) FROM bookings").fetchone()[0])
        log_count = int(conn.execute("SELECT COUNT(*) FROM process_logs").fetchone()[0])
        conn.execute("DELETE FROM bookings")
        conn.execute("DELETE FROM process_logs")
    return {"bookings": booking_count, "logs": log_count}


def save_process_log(stats: dict[str, Any]) -> None:
    init_db()
    with db_connect() as conn:
        conn.execute(
            """
            INSERT INTO process_logs (
                processed_at, report_date, report_month, output_file,
                raw_rows, included_rows, excluded_rows, ayo_rows, walk_in_rows,
                total_revenue, ayo_revenue, walk_in_revenue,
                source_file, feature, period_label
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                datetime.now().isoformat(timespec="seconds"),
                stats.get("selected_date") or None,
                stats.get("month", ""),
                stats.get("filename", ""),
                int(stats.get("raw_rows", 0)),
                int(stats.get("included_rows", 0)),
                int(stats.get("excluded_rows", 0)),
                int(stats.get("ayo_rows", 0)),
                int(stats.get("walk_in_rows", 0)),
                float(stats.get("total_revenue", 0)),
                float(stats.get("ayo_revenue", 0)),
                float(stats.get("walk_in_revenue", 0)),
                stats.get("source_file", ""),
                stats.get("feature", ""),
                stats.get("period_label", ""),
            ),
        )


def load_process_logs(limit: int = 20) -> list[dict[str, Any]]:
    init_db()
    with db_connect() as conn:
        rows = conn.execute(
            """
            SELECT *
            FROM process_logs
            ORDER BY processed_at DESC, id DESC
            LIMIT ?
            """,
            (limit,),
        ).fetchall()
    return [dict(row) for row in rows]


def latest_year_for_month(month: int | None = None) -> int | None:
    init_db()
    query = "SELECT MAX(booking_year) FROM bookings"
    params: list[Any] = []
    if month:
        query += " WHERE booking_month = ?"
        params.append(month)
    with db_connect() as conn:
        value = conn.execute(query, params).fetchone()[0]
    return int(value) if value else None


def infer_year(records: list[dict[str, Any]], month: int | None = None) -> int:
    years = []
    for record in records:
        booking_date = parse_date(record.get("Date of Booking"))
        if booking_date and (month is None or booking_date.month == month):
            years.append(booking_date.year)
    if years:
        return max(years)

    db_year = latest_year_for_month(month)
    if db_year:
        return db_year
    return datetime.now().year


def parse_compact_period(
    day_value: str | None,
    month_value: str | None,
    year_value: str | None,
    records_for_year: list[dict[str, Any]],
) -> tuple[date | None, date | None]:
    day_text = normalize_header(day_value)
    month_text = normalize_header(month_value)
    year_text = normalize_header(year_value)

    day = int(day_text) if day_text else None
    month = int(month_text) if month_text else None
    year = int(year_text) if year_text else infer_year(records_for_year, month)

    if day is not None and not 1 <= day <= 31:
        raise ValueError("Tanggal booking harus angka 1 sampai 31.")
    if month is not None and not 1 <= month <= 12:
        raise ValueError("Bulan rekap harus angka 1 sampai 12.")
    if day is not None and month is None:
        raise ValueError("Kalau tanggal diisi, bulan rekap juga harus dipilih.")

    report_month = date(year, month, 1) if month else None
    report_date = date(year, month, day) if day and month else None
    return report_date, report_month


def find_header_row(ws) -> int:
    seen_keys: set[str] = set()
    workbook_keys = {as_key(sheet_name) for sheet_name in getattr(ws.parent, "sheetnames", [])}

    def looks_like_generated_rekap() -> bool:
        keys = seen_keys | workbook_keys
        return (
            any(key.startswith("summary lapangan") for key in keys)
            or any(key.startswith("penjualan per kategori") for key in keys)
            or any(key.startswith("omset sewa lapangan") for key in keys)
            or any(key.startswith("omset keseluruhan") for key in keys)
            or ({"date of booking", "ayo", "walk in", "total"} <= keys)
        )

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
        values = [as_key(value) for value in row]
        seen_keys.update(value for value in values if value)
        if looks_like_generated_rekap():
            raise ValueError(
                "File ini sudah berupa workbook hasil rekap, bukan Excel mentah booking. "
                "Pilih file export booking mentah yang punya kolom 'Booking ID' dan 'Payment Method'."
            )
        if "booking id" in values and "payment method" in values:
            return row_idx
    # Bantu user kalau salah modul: file Olsera ('item group'/'order no') masuk ke Omset Lapangan.
    if seen_keys & {"item group", "order no", "order date", "sales name"}:
        raise ValueError(
            "File ini sepertinya export Olsera (ada kolom 'item group'/'order no'), bukan data booking "
            "lapangan. Pakai modul 'Omset Perkategori Olsera' untuk file ini."
        )
    raise ValueError("Header tidak ditemukan. File harus punya kolom Booking ID dan Payment Method.")


def extract_records(file_obj, should_cancel=None) -> dict[str, Any]:
    ensure_not_cancelled(should_cancel)
    wb = load_workbook(file_obj, data_only=True, read_only=True)
    ws = wb.active
    header_row = find_header_row(ws)

    title = normalize_header(ws["A1"].value) or "BC Padel Club"
    period = normalize_header(ws["A2"].value)
    header_values = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    raw_headers = [normalize_header(value) for value in header_values]
    headers = [h for h in raw_headers if h]
    header_positions = {
        normalize_header(value): index
        for index, value in enumerate(header_values)
        if normalize_header(value)
    }

    records: list[dict[str, Any]] = []
    for row_idx, row_values in enumerate(
        ws.iter_rows(min_row=header_row + 1, values_only=True),
        start=header_row + 1,
    ):
        if row_idx % 100 == 0:
            ensure_not_cancelled(should_cancel)
        row_data: dict[str, Any] = {"_source_row": row_idx}
        has_content = False
        for header, col_idx in header_positions.items():
            value = row_values[col_idx] if col_idx < len(row_values) else None
            if value not in (None, ""):
                has_content = True
            row_data[header] = format_excel_value(value)

        booking_id = normalize_header(row_data.get("Booking ID"))
        if has_content and booking_id:
            records.append(row_data)

    dates = [d for d in (parse_date(record.get("Date of Booking")) for record in records) if d]
    if not dates:
        wb.close()
        raise ValueError("Kolom Date of Booking kosong atau tidak terbaca.")

    ensure_not_cancelled(should_cancel)
    first_date = min(dates)
    last_date = max(dates)
    wb.close()
    return {
        "title": title,
        "period": period,
        "headers": headers,
        "records": records,
        "first_date": first_date,
        "last_date": last_date,
    }


def find_column_position(ws, column_name: str, max_row: int = 20) -> tuple[int, int]:
    _, row_idx, col_idx = find_any_column_position(ws, [column_name], max_row=max_row)
    return row_idx, col_idx


def find_any_column_position(
    ws,
    column_names: list[str] | tuple[str, ...],
    max_row: int = 20,
) -> tuple[str, int, int]:
    targets = {as_key(column_name): column_name for column_name in column_names}
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_row, values_only=True), start=1):
        for col_idx, value in enumerate(row):
            key = as_key(value)
            if key in targets:
                return normalize_header(value), row_idx, col_idx
    expected = "', '".join(column_names)
    raise ValueError(f"Kolom '{expected}' tidak ditemukan pada {max_row} baris pertama.")


def find_optional_column_position(
    ws,
    column_names: list[str] | tuple[str, ...],
    max_row: int = 20,
) -> tuple[str, int, int] | None:
    try:
        return find_any_column_position(ws, column_names, max_row=max_row)
    except ValueError:
        return None


ITEM_NAME_COLUMN_ALIASES = (
    "item name",
    "nama item",
    "product name",
    "nama produk",
    "produk",
)
SEWA_RAKET_GROUP_KEY = "sewa raket"
SEWA_RAKET_PREMIUM_CATEGORY = "SEWA RAKET - RAKET PREMIUM"
SEWA_RAKET_STANDAR_CATEGORY = "SEWA RAKET - RAKET STANDAR"


def category_from_group_and_item(group_value: Any, item_name_value: Any = None) -> str:
    """Kategori Olsera, dengan pecahan khusus SEWA RAKET berdasarkan item name."""
    group = normalize_header(group_value)
    if as_key(group) != SEWA_RAKET_GROUP_KEY:
        return group

    item_key = as_key(item_name_value)
    if "premium" in item_key:
        return SEWA_RAKET_PREMIUM_CATEGORY
    if "standar" in item_key or "standard" in item_key:
        return SEWA_RAKET_STANDAR_CATEGORY
    return group


def extract_categories(file_obj, filename: str, should_cancel=None) -> dict[str, Any]:
    ensure_not_cancelled(should_cancel)
    wb = load_workbook(file_obj, data_only=True, read_only=True)
    ws = wb.active
    try:
        group_header, header_row, group_col_idx = find_any_column_position(
            ws,
            ("item group", "group", "kategori"),
        )
    except ValueError as exc:
        wb.close()
        raise ValueError(f"{filename}: {exc}") from exc

    amount_column = find_optional_column_position(ws, ("amount", "total", "total amount"))
    date_column = find_optional_column_position(ws, ("order date", "tanggal", "date"))
    item_name_column = find_optional_column_position(ws, ITEM_NAME_COLUMN_ALIASES)
    amount_col_idx = amount_column[2] if amount_column else None
    date_col_idx = date_column[2] if date_column else None
    item_name_col_idx = item_name_column[2] if item_name_column else None
    counts: dict[str, int] = {}
    amounts: dict[str, float] = {}
    source_groups: dict[str, int] = {}
    dates: list[date] = []
    for row_idx, row_values in enumerate(
        ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1
    ):
        if row_idx % 500 == 0:
            ensure_not_cancelled(should_cancel)
        value = row_values[group_col_idx] if group_col_idx < len(row_values) else None
        group = normalize_header(value)
        if group:
            source_groups[group] = source_groups.get(group, 0) + 1
            item_name = (
                row_values[item_name_col_idx]
                if item_name_col_idx is not None and item_name_col_idx < len(row_values)
                else None
            )
            category = category_from_group_and_item(group, item_name)
            counts[category] = counts.get(category, 0) + 1
            if amount_col_idx is not None:
                amount_value = row_values[amount_col_idx] if amount_col_idx < len(row_values) else None
                amounts[category] = amounts.get(category, 0) + number_value(amount_value)
            if date_col_idx is not None:
                date_value = row_values[date_col_idx] if date_col_idx < len(row_values) else None
                parsed_date = parse_date(date_value)
                if parsed_date:
                    dates.append(parsed_date)

    wb.close()
    if not counts:
        raise ValueError(f"{filename}: Tidak ada data group/kategori yang terbaca pada kolom '{group_header}'.")

    return {
        "filename": filename,
        "group_column": group_header,
        "categories": counts,
        "source_groups": source_groups,
        "amounts": amounts,
        "total_rows": sum(counts.values()),
        "total_amount": sum(amounts.values()),
        "first_date": min(dates).isoformat() if dates else "",
        "last_date": max(dates).isoformat() if dates else "",
    }


def category_result_period(result: dict[str, Any]) -> tuple[date, date] | None:
    first_date = parse_date(result.get("first_date"))
    last_date = parse_date(result.get("last_date"))
    if not first_date and not last_date:
        return None
    first_date = first_date or last_date
    last_date = last_date or first_date
    if first_date is None or last_date is None:
        return None
    return (min(first_date, last_date), max(first_date, last_date))


def category_periods_overlap(
    first: tuple[date, date] | None,
    second: tuple[date, date] | None,
) -> bool:
    # Jika file tidak punya kolom tanggal, anggap berpotensi duplikat agar aman.
    if first is None or second is None:
        return True
    return first[0] <= second[1] and second[0] <= first[1]


def find_duplicate_categories(file_results: list[dict[str, Any]]) -> dict[str, list[str]]:
    owners: dict[str, list[tuple[str, tuple[date, date] | None]]] = {}
    display_names: dict[str, str] = {}
    for result in file_results:
        period = category_result_period(result)
        for category in result["categories"]:
            key = category.upper()
            entries = owners.setdefault(key, [])
            filename = str(result["filename"])
            if not any(existing_filename == filename for existing_filename, _ in entries):
                entries.append((filename, period))
            display_names.setdefault(key, category)

    duplicates: dict[str, list[str]] = {}
    for key, entries in owners.items():
        duplicate_indexes: set[int] = set()
        for first_index, (_, first_period) in enumerate(entries):
            for second_index in range(first_index + 1, len(entries)):
                if category_periods_overlap(first_period, entries[second_index][1]):
                    duplicate_indexes.update((first_index, second_index))
        if duplicate_indexes:
            duplicates[display_names[key]] = [
                filename for index, (filename, _) in enumerate(entries) if index in duplicate_indexes
            ]
    return duplicates


def duplicate_category_files_to_remove(
    file_results: list[dict[str, Any]],
) -> list[tuple[str, str]]:
    """File duplikat setelah file pertama, sesuai urutan pilihan pengguna.

    Hasil berisi pasangan ``(_source_path, filename)``. Evaluasi dilakukan
    bertahap supaya satu file yang dibuang karena group A tidak menyebabkan
    file lain ikut terbuang padahal masih diperlukan untuk group B.
    """
    # File yang valid (satu source group) diprioritaskan. Ini mencegah workbook
    # hasil rekap yang tidak sengaja dipilih mengalahkan file export mentah.
    ordered_results = sorted(
        enumerate(file_results),
        key=lambda item: (
            len(item[1].get("source_groups") or item[1].get("categories") or {}) > 1,
            item[0],
        ),
    )
    kept: list[dict[str, Any]] = []
    removed: list[tuple[str, str]] = []
    for _, result in ordered_results:
        if find_duplicate_categories([*kept, result]):
            removed.append(
                (str(result.get("_source_path") or ""), str(result.get("filename") or ""))
            )
        else:
            kept.append(result)
    return removed


def booking_channel(record: dict[str, Any]) -> str:
    booking_id = normalize_header(record.get("Booking ID")).upper()
    if booking_id.startswith("MN/"):
        return "Walk In"
    if booking_id.startswith("BK/"):
        return "AYO"

    payment_method = normalize_header(record.get("Payment Method")).upper()
    if "MANUAL" in payment_method or "WALK" in payment_method:
        return "Walk In"
    return "AYO"


def is_walk_in(record: dict[str, Any]) -> bool:
    return booking_channel(record) == "Walk In"


def should_exclude(record: dict[str, Any], exclude_cancelled: bool, exclude_failed: bool) -> bool:
    booking_id = normalize_header(record.get("Booking ID"))
    if booking_id in DEFAULT_EXCLUDED_BOOKING_IDS or record.get("_excluded_reason"):
        return True

    searchable = " ".join(normalize_header(record.get(header)) for header in IMPORTANT_HEADERS)
    upper = searchable.upper()
    if exclude_cancelled and any(token in upper for token in ("CANCEL", "BATAL")):
        return True
    if exclude_failed and any(token in upper for token in ("FAILED", "FAIL", "GAGAL")):
        return True
    return False


def split_records(
    records: list[dict[str, Any]], exclude_cancelled: bool, exclude_failed: bool
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    included: list[dict[str, Any]] = []
    excluded: list[dict[str, Any]] = []
    walk_in: list[dict[str, Any]] = []
    ayo: list[dict[str, Any]] = []

    for record in records:
        if should_exclude(record, exclude_cancelled, exclude_failed):
            excluded.append(record)
            continue
        included.append(record)
        if is_walk_in(record):
            walk_in.append(record)
        else:
            ayo.append(record)
    return included, walk_in, ayo, excluded


def find_duplicate_booking_ids(records: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    display_ids: dict[str, str] = {}
    for record in records:
        booking_id = normalize_header(record.get("Booking ID"))
        if not booking_id:
            continue
        key = booking_id.upper()
        display_ids.setdefault(key, booking_id)
        booking_date = parse_date(record.get("Date of Booking"))
        start_time = parse_time_value(record.get("Booking Period Start Time"))
        grouped[key].append(
            {
                "booking_id": booking_id,
                "date": booking_date.isoformat() if booking_date else normalize_header(record.get("Date of Booking")),
                "court": normalize_header(record.get("Court")),
                "start_time": start_time.strftime("%H:%M:%S") if start_time else normalize_header(record.get("Booking Period Start Time")),
                "revenue": number_value(record.get("Revenue Venue")),
                "source_row": normalize_header(record.get("_source_row") or record.get("No")),
            }
        )

    return {
        display_ids[key]: values
        for key, values in grouped.items()
        if len(values) > 1
    }


def court_sort_key(record_or_name: dict[str, Any] | str) -> tuple[int, int, str]:
    court = record_or_name if isinstance(record_or_name, str) else normalize_header(record_or_name.get("Court"))
    match = re.search(r"(\d+)", court)
    number = int(match.group(1)) if match else 999
    kind = 2 if "PICKLE" in court.upper() else 1
    return kind, number, court


def display_court_label(court: str) -> str:
    normalized = normalize_header(court)
    if "PICKLE" in normalized.upper():
        return "Pickle Court"
    match = re.search(r"(\d+)", normalized)
    if match:
        return f"Court {int(match.group(1))}"
    return normalized


def channel_sort_key(record: dict[str, Any]) -> int:
    channel = booking_channel(record)
    return 0 if channel == "AYO" else 1


def record_sort_key(record: dict[str, Any]) -> tuple[Any, ...]:
    booking_date = parse_date(record.get("Date of Booking")) or date.max
    start_time = parse_time_value(record.get("Booking Period Start Time")) or time.max
    return court_sort_key(record) + (channel_sort_key(record), booking_date, start_time, normalize_header(record.get("Booking ID")))


def record_date_sort_key(record: dict[str, Any]) -> tuple[Any, ...]:
    booking_date = parse_date(record.get("Date of Booking")) or date.max
    start_time = parse_time_value(record.get("Booking Period Start Time")) or time.max
    original_no = number_value(record.get("No")) or 999999
    return (
        booking_date,
        start_time,
        original_no,
    ) + court_sort_key(record) + (
        channel_sort_key(record),
        normalize_header(record.get("Booking ID")),
    )


def court_time_sort_key(record: dict[str, Any]) -> tuple[Any, ...]:
    start_time = parse_time_value(record.get("Booking Period Start Time")) or time.max
    original_no = number_value(record.get("No")) or 999999
    return court_sort_key(record) + (
        start_time,
        original_no,
        channel_sort_key(record),
        normalize_header(record.get("Booking ID")),
    )


def month_label(first_date: date) -> tuple[str, str, str]:
    month = MONTH_NAMES[first_date.month]
    return month, f"{month}'{str(first_date.year)[-2:]}", f"{month} {first_date.year}"


def setup_sheet(ws, tab_color: str = "1F6F78") -> None:
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = tab_color


def style_title(ws, title: str, period: str, last_col: int) -> None:
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="174E57")
    ws["A1"].alignment = Alignment(vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, last_col))
    ws["A2"] = period
    ws["A2"].font = Font(bold=True, color="1F2937")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(1, last_col))


def style_header(ws, row: int, last_col: int) -> None:
    fill = PatternFill("solid", fgColor="0F766E")
    border = Border(bottom=Side(style="thin", color="94A3B8"))
    for col in range(1, last_col + 1):
        cell = ws.cell(row, col)
        cell.fill = fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border


def apply_table_borders(
    ws,
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
    color: str = "000000",
) -> None:
    if max_row < min_row or max_col < min_col:
        return
    side = Side(style="thin", color=color)
    border = Border(left=side, right=side, top=side, bottom=side)
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            ws.cell(row, col).border = border


def set_widths(ws, headers: list[str], start_col: int = 1) -> None:
    width_by_header = {
        "No": 8,
        "Tanggal": 14,
        "Booking ID": 24,
        "Venue": 18,
        "Court": 20,
        "Court ID": 10,
        "Sports Category": 16,
        "Customer Name": 24,
        "Customer Email": 28,
        "Customer Phone": 18,
        "Username Customer": 20,
        "Date of Booking": 16,
        "Booking Period Start Time": 16,
        "Booking Period End Time": 16,
        "Session Length": 16,
        "Payment ID": 22,
        "Payment Method": 18,
        "Status": 14,
    }
    for offset, header in enumerate(headers):
        col_letter = get_column_letter(start_col + offset)
        if header in MONEY_HEADERS:
            width = 16
        else:
            width = width_by_header.get(header, min(max(len(header) + 3, 12), 28))
        ws.column_dimensions[col_letter].width = width


def write_table_sheet(
    wb: Workbook,
    sheet_name: str,
    title: str,
    period: str,
    headers: list[str],
    records: list[dict[str, Any]],
    tab_color: str,
    total_label_col: int | None = None,
    group_by_date: bool = False,
) -> None:
    ws = wb.create_sheet(sheet_name)
    setup_sheet(ws, tab_color)
    style_title(ws, title, period, len(headers))

    header_row = 4
    for col, header in enumerate(headers, 1):
        ws.cell(header_row, col, header)
    style_header(ws, header_row, len(headers))

    start_row = header_row + 1
    row_idx = start_row

    if group_by_date:
        grouped: dict[date, list[dict[str, Any]]] = defaultdict(list)
        for record in sorted(records, key=record_date_sort_key):
            booking_date = parse_date(record.get("Date of Booking")) or date.max
            grouped[booking_date].append(record)

        running_no = 1
        for booking_date in sorted(grouped):
            group = grouped[booking_date]
            group_start = row_idx
            for record in group:
                for col, header in enumerate(headers, 1):
                    if header == "No":
                        value = record.get(header) or running_no
                    else:
                        value = record.get(header)
                    ws.cell(row_idx, col, value)
                running_no += 1
                row_idx += 1

            subtotal_row = row_idx
            for col, header in enumerate(headers, 1):
                if header in MONEY_HEADERS:
                    letter = get_column_letter(col)
                    ws.cell(subtotal_row, col, f"=SUM({letter}{group_start}:{letter}{subtotal_row - 1})")
            for col in range(1, len(headers) + 1):
                cell = ws.cell(subtotal_row, col)
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="FFFF00")
            row_idx += 1

        last_row = max(row_idx - 1, header_row)
    else:
        for row_offset, record in enumerate(records):
            row_idx = start_row + row_offset
            for col, header in enumerate(headers, 1):
                value = row_offset + 1 if header == "No" else record.get(header)
                ws.cell(row_idx, col, value)

        total_row = start_row + len(records)
        if records:
            if total_label_col:
                ws.cell(total_row, total_label_col, "TOTAL")
            for col, header in enumerate(headers, 1):
                if header in MONEY_HEADERS:
                    letter = get_column_letter(col)
                    ws.cell(total_row, col, f"=SUM({letter}{start_row}:{letter}{total_row - 1})")
            for col in range(1, len(headers) + 1):
                cell = ws.cell(total_row, col)
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="E2E8F0")
        last_row = max(total_row, header_row)

    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{last_row}"
    format_body(ws, header_row + 1, last_row, headers)
    set_widths(ws, headers)


def format_body(ws, start_row: int, end_row: int, headers: list[str]) -> None:
    if end_row < start_row:
        return

    thin = Side(style="thin", color="E2E8F0")
    border = Border(bottom=thin)
    alignment = Alignment(vertical="top", wrap_text=False)
    money_format = '#,##0'

    template_styles = []
    for col, header in enumerate(headers, 1):
        cell = ws.cell(start_row, col)
        cell.border = border
        cell.alignment = alignment
        if header in MONEY_HEADERS:
            cell.number_format = money_format
        if "Date" in header or header == "Tanggal":
            cell.number_format = "yyyy-mm-dd"
        template_styles.append(copy(cell._style))

    style_cache = {}
    for row in range(start_row + 1, end_row + 1):
        for col, template_style in enumerate(template_styles, 1):
            cell = ws.cell(row, col)
            if cell.has_style and cell._style is not None:
                cache_key = (
                    tuple(cell._style),
                    template_style.borderId,
                    template_style.alignmentId,
                    template_style.numFmtId,
                )
                style = style_cache.get(cache_key)
                if style is None:
                    # Preserve existing bold/fill subtotal styles while applying shared body formats.
                    style = copy(cell._style)
                    style.borderId = template_style.borderId
                    style.alignmentId = template_style.alignmentId
                    style.numFmtId = template_style.numFmtId
                    style_cache[cache_key] = style
                cell._style = style
            else:
                cell._style = template_style


def all_headers(raw_headers: list[str]) -> list[str]:
    headers = ["Tanggal"]
    for header in raw_headers:
        if header not in {"No", "Date of Booking"}:
            headers.append(header)
    return headers


def to_all_record(record: dict[str, Any], headers: list[str]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for header in headers:
        if header == "Tanggal":
            result[header] = record.get("Date of Booking")
        else:
            result[header] = record.get(header)
    return result


def write_all_sheet(
    wb: Workbook,
    sheet_name: str,
    period: str,
    raw_headers: list[str],
    records: list[dict[str, Any]],
    group_by_date: bool = False,
) -> None:
    headers = all_headers(raw_headers)
    ws = wb.create_sheet(sheet_name)
    setup_sheet(ws, "334155")
    style_title(ws, "OMSET SEWA LAPANGAN (ALL)", period, len(headers))
    header_row = 4
    for col, header in enumerate(headers, 1):
        ws.cell(header_row, col, header)
    style_header(ws, header_row, len(headers))

    row_idx = header_row + 1
    if group_by_date:
        grouped_by_date: dict[date, list[dict[str, Any]]] = defaultdict(list)
        for record in sorted(records, key=record_date_sort_key):
            booking_date = parse_date(record.get("Date of Booking")) or date.max
            grouped_by_date[booking_date].append(record)

        for booking_date in sorted(grouped_by_date):
            date_records = grouped_by_date[booking_date]
            grouped_by_court: dict[str, list[dict[str, Any]]] = defaultdict(list)
            for record in date_records:
                grouped_by_court[normalize_header(record.get("Court"))].append(record)

            daily_subtotal_rows: list[int] = []
            wrote_date = False
            for court in sorted(grouped_by_court, key=court_sort_key):
                group_start = row_idx
                for record in sorted(grouped_by_court[court], key=court_time_sort_key):
                    all_record = to_all_record(record, headers)
                    all_record["Tanggal"] = booking_date if not wrote_date else None
                    wrote_date = True
                    for col, header in enumerate(headers, 1):
                        ws.cell(row_idx, col, all_record.get(header))
                    row_idx += 1

                subtotal_row = row_idx
                for col, header in enumerate(headers, 1):
                    if header in MONEY_HEADERS:
                        letter = get_column_letter(col)
                        ws.cell(subtotal_row, col, f"=SUM({letter}{group_start}:{letter}{subtotal_row - 1})")
                for col in range(1, len(headers) + 1):
                    ws.cell(subtotal_row, col).font = Font(bold=True)
                    ws.cell(subtotal_row, col).fill = PatternFill("solid", fgColor="FFFF00")
                daily_subtotal_rows.append(subtotal_row)
                row_idx += 1

            total_row = row_idx
            total_label_col = 2 if len(headers) > 1 else 1
            ws.cell(total_row, total_label_col, "TOTAL")
            for col, header in enumerate(headers, 1):
                if header in MONEY_HEADERS and daily_subtotal_rows:
                    letter = get_column_letter(col)
                    refs = "+".join(f"{letter}{r}" for r in daily_subtotal_rows)
                    ws.cell(total_row, col, f"={refs}")
            for col in range(1, len(headers) + 1):
                ws.cell(total_row, col).font = Font(bold=True)
                ws.cell(total_row, col).fill = PatternFill("solid", fgColor="FFFF00")
            row_idx += 1
    else:
        grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for record in sorted(records, key=record_sort_key):
            grouped[normalize_header(record.get("Court"))].append(record)

        grand_total_rows: list[int] = []
        for court in sorted(grouped, key=court_sort_key):
            group_start = row_idx
            for record in grouped[court]:
                all_record = to_all_record(record, headers)
                for col, header in enumerate(headers, 1):
                    ws.cell(row_idx, col, all_record.get(header))
                row_idx += 1

            subtotal_row = row_idx
            ws.cell(subtotal_row, 1, f"Subtotal {court}")
            for col, header in enumerate(headers, 1):
                if header in MONEY_HEADERS:
                    letter = get_column_letter(col)
                    ws.cell(subtotal_row, col, f"=SUM({letter}{group_start}:{letter}{subtotal_row - 1})")
            for col in range(1, len(headers) + 1):
                ws.cell(subtotal_row, col).font = Font(bold=True)
                ws.cell(subtotal_row, col).fill = PatternFill("solid", fgColor="F1F5F9")
            grand_total_rows.append(subtotal_row)
            row_idx += 1

        total_row = row_idx
        ws.cell(total_row, 1, "TOTAL")
        for col, header in enumerate(headers, 1):
            if header in MONEY_HEADERS and grand_total_rows:
                letter = get_column_letter(col)
                refs = ",".join(f"{letter}{r}" for r in grand_total_rows)
                ws.cell(total_row, col, f"=SUM({refs})")
        for col in range(1, len(headers) + 1):
            ws.cell(total_row, col).font = Font(bold=True, color="FFFFFF")
            ws.cell(total_row, col).fill = PatternFill("solid", fgColor="174E57")

    last_row = max(row_idx - 1, header_row)
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{last_row}"
    format_body(ws, header_row + 1, last_row, headers)
    set_widths(ws, headers)


def write_summary_sheet(
    wb: Workbook,
    sheet_name: str,
    first_date: date,
    records: list[dict[str, Any]],
    summary_dates: list[date] | None = None,
) -> None:
    ws = wb.create_sheet(sheet_name, 0)
    ws.sheet_properties.tabColor = "0F766E"
    ws.sheet_view.showGridLines = False
    month_name, _, month_year = month_label(first_date)
    last_day = calendar.monthrange(first_date.year, first_date.month)[1]
    days = summary_dates or [date(first_date.year, first_date.month, day) for day in range(1, last_day + 1)]
    courts = sorted({normalize_header(record.get("Court")) for record in records if record.get("Court")}, key=court_sort_key)

    ws["A1"] = "SUMMARY LAPANGAN PADEL & PICKLE"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="174E57")
    ws["A2"] = f"Bulan: {month_year.upper()}"

    header_row_1 = 4
    header_row_2 = 5
    ws.cell(header_row_1, 1, "Date of \nBooking")
    ws.merge_cells(start_row=header_row_1, start_column=1, end_row=header_row_2, end_column=1)

    col = 2
    court_columns: list[tuple[str, int, int]] = []
    for court in courts:
        ws.cell(header_row_1, col, display_court_label(court))
        ws.merge_cells(start_row=header_row_1, start_column=col, end_row=header_row_1, end_column=col + 1)
        ws.cell(header_row_2, col, "AYO")
        ws.cell(header_row_2, col + 1, "Walk In")
        court_columns.append((court, col, col + 1))
        col += 2
    total_col = col
    ws.cell(header_row_1, total_col, "TOTAL")
    ws.merge_cells(start_row=header_row_1, start_column=total_col, end_row=header_row_2, end_column=total_col)

    style_header(ws, header_row_1, total_col)
    style_header(ws, header_row_2, total_col)
    ws.freeze_panes = "B6"

    summary = defaultdict(float)
    for record in records:
        booking_date = parse_date(record.get("Date of Booking"))
        court = normalize_header(record.get("Court"))
        if not booking_date or not court:
            continue
        channel = booking_channel(record)
        summary[(booking_date, court, channel)] += number_value(record.get("Revenue Venue"))

    start_row = 6
    for row_offset, current_day in enumerate(days):
        row_idx = start_row + row_offset
        ws.cell(row_idx, 1, current_day)
        col = 2
        for court in courts:
            ws.cell(row_idx, col, summary[(current_day, court, "AYO")] or 0)
            ws.cell(row_idx, col + 1, summary[(current_day, court, "Walk In")] or 0)
            col += 2
        first_amount_col = get_column_letter(2)
        last_amount_col = get_column_letter(total_col - 1)
        ws.cell(row_idx, total_col, f"=SUM({first_amount_col}{row_idx}:{last_amount_col}{row_idx})")

    total_row = start_row + len(days)
    ws.cell(total_row, 1, "TOTAL")
    for amount_col in range(2, total_col + 1):
        letter = get_column_letter(amount_col)
        ws.cell(total_row, amount_col, f"=SUM({letter}{start_row}:{letter}{total_row - 1})")
        ws.cell(total_row, amount_col).font = Font(bold=True)
    ws.cell(total_row, 1).font = Font(bold=True)
    for col_idx in range(1, total_col + 1):
        ws.cell(total_row, col_idx).fill = PatternFill("solid", fgColor="E2E8F0")

    notes_start = total_row + 2
    footer_label_col = max(1, total_col - 5)
    padel_col = footer_label_col + 1
    pickle_col = footer_label_col + 2
    footer_total_col = footer_label_col + 3
    diff_col = footer_label_col + 4

    ws.cell(notes_start, padel_col, "PADEL")
    ws.cell(notes_start, pickle_col, "PICKLE")
    ws.cell(notes_start, footer_total_col, "TOTAL")
    ws.cell(notes_start, diff_col, "Selisih OLSERA - AYO")
    ws.cell(notes_start + 1, footer_label_col, "AYO")
    ws.cell(notes_start + 2, footer_label_col, "WALK IN")
    ws.cell(notes_start + 3, footer_label_col, "TOTAL")
    ws.cell(notes_start + 4, footer_label_col, "Average")

    def add_refs(refs: list[str]) -> str | int:
        return "=" + "+".join(refs) if refs else 0

    padel_ayo_refs: list[str] = []
    padel_walk_refs: list[str] = []
    pickle_ayo_refs: list[str] = []
    pickle_walk_refs: list[str] = []
    for court, ayo_col, walk_col in court_columns:
        ayo_ref = f"{get_column_letter(ayo_col)}{total_row}"
        walk_ref = f"{get_column_letter(walk_col)}{total_row}"
        if "PICKLE" in court.upper():
            pickle_ayo_refs.append(ayo_ref)
            pickle_walk_refs.append(walk_ref)
        else:
            padel_ayo_refs.append(ayo_ref)
            padel_walk_refs.append(walk_ref)

    ws.cell(notes_start + 1, padel_col, add_refs(padel_ayo_refs))
    ws.cell(notes_start + 1, pickle_col, add_refs(pickle_ayo_refs))
    ws.cell(notes_start + 1, footer_total_col, f"=SUM({get_column_letter(padel_col)}{notes_start + 1}:{get_column_letter(pickle_col)}{notes_start + 1})")
    ws.cell(notes_start + 2, padel_col, add_refs(padel_walk_refs))
    ws.cell(notes_start + 2, pickle_col, add_refs(pickle_walk_refs))
    ws.cell(notes_start + 2, footer_total_col, f"=SUM({get_column_letter(padel_col)}{notes_start + 2}:{get_column_letter(pickle_col)}{notes_start + 2})")
    ws.cell(notes_start + 3, padel_col, f"=SUM({get_column_letter(padel_col)}{notes_start + 1}:{get_column_letter(padel_col)}{notes_start + 2})")
    ws.cell(notes_start + 3, pickle_col, f"=SUM({get_column_letter(pickle_col)}{notes_start + 1}:{get_column_letter(pickle_col)}{notes_start + 2})")
    ws.cell(notes_start + 3, footer_total_col, f"=SUM({get_column_letter(padel_col)}{notes_start + 3}:{get_column_letter(pickle_col)}{notes_start + 3})")
    average_days = len(days) if summary_dates else last_day
    ws.cell(notes_start + 4, padel_col, f"={get_column_letter(padel_col)}{notes_start + 3}/{average_days}")
    ws.cell(notes_start + 4, pickle_col, f"={get_column_letter(pickle_col)}{notes_start + 3}/{average_days}")

    for col_idx in range(padel_col, min(diff_col, total_col) + 1):
        ws.cell(notes_start, col_idx).font = Font(bold=True, color="FFFFFF")
        ws.cell(notes_start, col_idx).fill = PatternFill("solid", fgColor="174E57")
        ws.cell(notes_start, col_idx).alignment = Alignment(horizontal="center")
    for row in range(notes_start + 1, notes_start + 5):
        ws.cell(row, footer_label_col).font = Font(bold=True)
        for col_idx in range(padel_col, min(footer_total_col, total_col) + 1):
            ws.cell(row, col_idx).number_format = '#,##0'
            if row == notes_start + 3:
                ws.cell(row, col_idx).font = Font(bold=True)

    for row in range(start_row, notes_start + 5):
        ws.cell(row, 1).number_format = "dd-mmm-yy"
        for col_idx in range(2, total_col + 1):
            ws.cell(row, col_idx).number_format = '#,##0'

    ws.column_dimensions["A"].width = 18
    for col_idx in range(2, total_col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15


def filter_records(
    records: list[dict[str, Any]],
    report_date: date | None,
    report_month: date | None,
    report_dates: list[date] | None = None,
) -> list[dict[str, Any]]:
    if report_dates:
        selected_dates = set(report_dates)
        return [record for record in records if parse_date(record.get("Date of Booking")) in selected_dates]

    if report_date:
        return [record for record in records if parse_date(record.get("Date of Booking")) == report_date]

    if report_month:
        return [
            record
            for record in records
            if (booking_date := parse_date(record.get("Date of Booking")))
            and booking_date.year == report_month.year
            and booking_date.month == report_month.month
        ]

    return list(records)


def choose_report_month(
    records: list[dict[str, Any]],
    report_date: date | None,
    report_month: date | None,
    report_dates: list[date] | None = None,
) -> date:
    if report_dates:
        return min(report_dates).replace(day=1)
    if report_month:
        return report_month.replace(day=1)
    if report_date:
        return report_date.replace(day=1)

    dates = [d for d in (parse_date(record.get("Date of Booking")) for record in records) if d]
    if not dates:
        raise ValueError("Tanggal booking tidak terbaca.")
    first = min(dates)
    return first.replace(day=1)


def db_recent_months(limit: int = 12) -> list[date]:
    init_db()
    with db_connect() as conn:
        rows = conn.execute(
            """
            SELECT DISTINCT booking_year, booking_month FROM bookings
            ORDER BY booking_year DESC, booking_month DESC LIMIT ?
            """,
            (limit,),
        ).fetchall()
    return [date(row["booking_year"], row["booking_month"], 1) for row in rows]


def detect_dates_in_records(records: list[dict[str, Any]]) -> list[date]:
    return sorted({d for d in (parse_date(r.get("Date of Booking")) for r in records) if d})


def detect_months_in_records(records: list[dict[str, Any]]) -> list[date]:
    return sorted({date(d.year, d.month, 1) for d in detect_dates_in_records(records)})


def safe_filename(name: str) -> str:
    cleaned = re.sub(r'[\\/:*?"<>|]', "", name).strip()
    return cleaned or "Omset Lapangan"


def build_omset_filename(
    report_date: date | None,
    report_month: date | None,
    report_dates: list[date] | None,
) -> str:
    """Nama file hasil rekap Omset Lapangan sesuai tanggal yang terdeteksi.

    - 1 tanggal  -> "Omset Lapangan 1 Juni 2026"
    - multi hari -> "Omset Lapangan 1-11 Juni 2026"
    - sebulan    -> "Omset Lapangan Juni 2026"
    """
    if report_dates:
        days = sorted(report_dates)
        first, last = days[0], days[-1]
        if first == last:
            name = f"Omset Lapangan {first.day} {MONTH_NAMES[first.month]} {first.year}"
        elif (first.month, first.year) == (last.month, last.year):
            name = f"Omset Lapangan {first.day}-{last.day} {MONTH_NAMES[first.month]} {first.year}"
        else:
            name = (
                f"Omset Lapangan {first.day} {MONTH_NAMES[first.month]} {first.year} - "
                f"{last.day} {MONTH_NAMES[last.month]} {last.year}"
            )
    elif report_date:
        name = f"Omset Lapangan {report_date.day} {MONTH_NAMES[report_date.month]} {report_date.year}"
    elif report_month:
        name = f"Omset Lapangan {MONTH_NAMES[report_month.month]} {report_month.year}"
    else:
        name = "Omset Lapangan"
    return safe_filename(name)


def unique_output_path(base_name: str, suffix: str = ".xlsx") -> Path:
    """Path di OUTPUT_DIR; kalau nama sudah ada, tambahkan ' (2)', ' (3)', dst."""
    candidate = OUTPUT_DIR / f"{base_name}{suffix}"
    counter = 2
    while candidate.exists():
        candidate = OUTPUT_DIR / f"{base_name} ({counter}){suffix}"
        counter += 1
    return candidate


def _safe_pdf_cell(value: Any) -> str:
    """Bersihkan teks PDF agar aman ditulis sebagai nilai sel Excel."""
    text = "" if value is None else str(value)
    text = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", text).strip()
    text = text[:32767]
    if text.startswith(("=", "+", "-", "@")):
        text = "'" + text
    return text


def _pdf_text_rows(text: str) -> list[list[str]]:
    rows: list[list[str]] = []
    for raw_line in (text or "").splitlines():
        line = raw_line.strip()
        if not line:
            continue
        cells = [_safe_pdf_cell(part) for part in re.split(r"\t+|\s{2,}", line)]
        rows.append([cell for cell in cells if cell] or [_safe_pdf_cell(line)])
    return rows


def _unique_sheet_title(source: str, used_titles: set[str]) -> str:
    cleaned = re.sub(r"[\\/*?:\[\]]", " ", source)
    cleaned = re.sub(r"\s+", " ", cleaned).strip() or "Halaman"
    cleaned = cleaned[:31]
    candidate = cleaned
    counter = 2
    while candidate.casefold() in used_titles:
        suffix = f" ({counter})"
        candidate = cleaned[: 31 - len(suffix)].rstrip() + suffix
        counter += 1
    used_titles.add(candidate.casefold())
    return candidate


def _style_pdf_sheet(ws, header_rows: set[int]) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    thin = Side(style="thin", color="D7E3E6")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)
            if cell.row in header_rows:
                cell.fill = PatternFill("solid", fgColor="0F4A5D")
                cell.font = Font(color="FFFFFF", bold=True)

    for column in range(1, ws.max_column + 1):
        values = [str(ws.cell(row, column).value or "") for row in range(1, min(ws.max_row, 150) + 1)]
        width = max((len(part) for value in values for part in value.splitlines()), default=0)
        ws.column_dimensions[get_column_letter(column)].width = min(45, max(12, width + 2))


# ── Rincian Penjualan harian (Olsera) ───────────────────────────────────────
# Layout 19 kolom (A-S) mengikuti contoh "Rincian penjualana Pelanggan.xlsx".
_RINCIAN_NCOLS = 19
_RINCIAN_MERGES_SUMMARY_A = [(1, 2), (3, 5), (6, 9), (10, 12), (13, 16), (17, 19)]
_RINCIAN_MERGES_SUMMARY_B = [(1, 2), (3, 5), (6, 9), (10, 12), (13, 19)]
_RINCIAN_MERGES_ROW = [(2, 3), (9, 10), (12, 13), (16, 17)]
_RINCIAN_MERGES_TOTAL = [(1, 7), (9, 10), (12, 13), (16, 17)]
_RINCIAN_COL_WIDTHS = {1: 16.8, 2: 12.1, 3: 15.8, 4: 17.4, 5: 0.1, 9: 9.0, 10: 9.8,
                       11: 9.0, 12: 9.0, 14: 11.2, 15: 9.0, 16: 0.1, 19: 11.0}
_RINCIAN_TEXT_COLS = {1, 2, 4, 5, 6, 7}  # rata kiri (No. Pesanan, Tanggal, Penjual, dst.)
_EN_TO_ID_MONTH = {
    "January": "Januari", "February": "Februari", "March": "Maret", "April": "April",
    "May": "Mei", "June": "Juni", "July": "Juli", "August": "Agustus",
    "September": "September", "October": "Oktober", "November": "November", "December": "Desember",
}


def _rincian_cell(value: Any, keep_newlines: bool = False) -> str:
    text = "" if value is None else str(value)
    text = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", text)
    if not keep_newlines:
        text = re.sub(r"\s*\n\s*", " ", text)
    text = text.strip()[:32767]
    if text.startswith(("=", "+", "@")):
        text = "'" + text
    return text


def _rincian_pad(row) -> list[Any]:
    cells = list(row or [])[:_RINCIAN_NCOLS]
    return cells + [None] * (_RINCIAN_NCOLS - len(cells))


# Posisi kolom (0-based) untuk 8 nilai uang di baris data/total:
# Total Penjualan, Pengiriman+Pajak, Modal, Laba, Biaya Layanan, Tambahan, Diskon, Jumlah Ditebus.
_RINCIAN_MONEY_COLS = [8, 10, 11, 13, 14, 15, 17, 18]


def _rincian_total_cells(cells: list[Any]) -> list[Any]:
    """Baris 'Total' di PDF sering punya jumlah kolom berbeda; tata ulang ke 19 kolom.

    Strukturnya: label + Qty + 8 nilai uang. Diambil berurutan dari sel non-kosong
    lalu ditempatkan ke posisi kolom yang sesuai layout (Qty di kolom H).
    """
    values = [c for c in cells[1:] if _rincian_cell(c)]
    out: list[Any] = [None] * _RINCIAN_NCOLS
    out[0] = cells[0]
    if values:
        out[7] = values[0]  # Qty
        for idx, value in enumerate(values[1:]):
            if idx < len(_RINCIAN_MONEY_COLS):
                out[_RINCIAN_MONEY_COLS[idx]] = value
    return out


def _is_rincian_penjualan(pdf) -> bool:
    try:
        text = (pdf.pages[0].extract_text() or "")[:300]
    except Exception:
        return False
    return "Rincian Penjualan" in text


def _rincian_merges(kind: str) -> list[tuple[int, int]]:
    return {
        "title": [(1, _RINCIAN_NCOLS)],
        "summaryA": _RINCIAN_MERGES_SUMMARY_A,
        "summaryB": _RINCIAN_MERGES_SUMMARY_B,
        "total": _RINCIAN_MERGES_TOTAL,
    }.get(kind, _RINCIAN_MERGES_ROW)


def _rincian_row_height(kind: str) -> float:
    return {"title": 37.5, "summaryA": 28.5, "summaryB": 31.5, "header": 25.5}.get(kind, 15.75)


def _write_rincian_title(ws, r: int, title_text: str, band_fill, border) -> None:
    """Baris judul: kiri = 'Rincian Penjualan' + periode, kanan = nama club besar."""
    lines = [ln.strip() for ln in (title_text or "").split("\n") if ln.strip()]
    first = lines[0] if lines else "Rincian Penjualan"
    periode = next((ln for ln in lines if ln.lower().startswith("periode")), "")
    club = first.replace("Rincian Penjualan", "").strip() or "BC PADEL CLUB"
    left = "Rincian Penjualan" + (("\n" + periode) if periode else "")

    for c in range(1, _RINCIAN_NCOLS + 1):
        cell = ws.cell(row=r, column=c)
        cell.fill = band_fill
        cell.border = border

    left_cell = ws.cell(row=r, column=1, value=left)
    left_cell.font = Font(size=10)
    left_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    club_cell = ws.cell(row=r, column=6, value=club)
    club_cell.font = Font(size=16, bold=True)
    club_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=_RINCIAN_NCOLS)


def _write_rincian_penjualan_sheet(workbook, pdf, path, used_titles, should_cancel=None):
    """Tulis 1 sheet bergaya laporan 'Rincian Penjualan' harian dari sebuah PDF."""
    title_text = ""
    layout: list[tuple[str, list[Any]]] = []  # (kind, 19 sel)
    data_count = 0

    for page in pdf.pages:
        ensure_not_cancelled(should_cancel)
        tables = page.extract_tables() or []
        if not tables:
            continue
        for raw in tables[0]:
            cells = _rincian_pad(raw)
            head = _rincian_cell(cells[0])
            if head.startswith("Rincian Penjualan"):
                title_text = _rincian_cell(cells[0], keep_newlines=True)
                layout.append(("title", cells))
            elif head.startswith("Total Penjualan"):
                layout.append(("summaryA", cells))
            elif head.startswith("Diskon") and not any(k == "summaryB" for k, _ in layout):
                layout.append(("summaryB", cells))
            elif head == "No. Pesanan":
                if not any(k == "header" for k, _ in layout):
                    layout.append(("header", cells))
            elif head.startswith("Total -"):
                layout.append(("total", _rincian_total_cells(cells)))
            elif re.match(r"^[A-Z0-9]{8,}$", head):
                layout.append(("data", cells))
                data_count += 1

    # Nama sheet = tanggal periode (mis. "01 Juni"), fallback ke nama file.
    sheet_name = path.stem
    match = re.search(r"Periode\s+(\d{1,2})\s+([A-Za-z]+)\s+(\d{4})", title_text)
    if match:
        sheet_name = f"{int(match.group(1)):02d} {_EN_TO_ID_MONTH.get(match.group(2), match.group(2))}"
    title = _unique_sheet_title(sheet_name, used_titles)
    ws = workbook.create_sheet(title)

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    band_fill = PatternFill("solid", fgColor="9DC3E6")    # band biru: judul & total
    header_fill = PatternFill("solid", fgColor="BFBFBF")  # header abu-abu

    for r, (kind, cells) in enumerate(layout, start=1):
        if kind == "title":
            _write_rincian_title(ws, r, title_text, band_fill, border)
            ws.row_dimensions[r].height = _rincian_row_height(kind)
            continue
        for c in range(1, _RINCIAN_NCOLS + 1):
            keep_nl = kind in ("summaryA", "summaryB") or (kind == "data" and c == 2)
            cell = ws.cell(row=r, column=c, value=_rincian_cell(cells[c - 1], keep_newlines=keep_nl) or None)
            cell.border = border
            cell.font = Font(size=10, bold=(kind in ("header", "total")))
            if kind == "header":
                cell.fill = header_fill
            elif kind == "total":
                cell.fill = band_fill
            if kind == "total" and c == 1:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            elif kind == "header" or (kind not in ("summaryA", "summaryB") and c not in _RINCIAN_TEXT_COLS):
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        for c1, c2 in _rincian_merges(kind):
            if c2 > c1:
                ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
        ws.row_dimensions[r].height = _rincian_row_height(kind)

    for col, width in _RINCIAN_COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.sheet_view.showGridLines = False

    return title, len(pdf.pages), data_count


def convert_pdfs_to_excel(
    pdf_paths: list[Path | str],
    output_path: Path | None = None,
    should_cancel=None,
    progress=None,
) -> tuple[Path, dict[str, Any]]:
    """Ekspor tabel atau teks dari beberapa PDF ke satu workbook Excel."""
    try:
        import pdfplumber
    except ImportError as exc:
        raise RuntimeError("Komponen pembaca PDF belum terpasang. Instal requirements.txt lalu coba lagi.") from exc

    paths = [Path(path) for path in pdf_paths]
    if not paths:
        raise ValueError("Pilih minimal satu file PDF.")
    for path in paths:
        if not path.is_file():
            raise ValueError(f"File PDF tidak ditemukan: {path}")
        if path.suffix.lower() != ".pdf":
            raise ValueError(f"Format file harus PDF: {path.name}")

    ensure_not_cancelled(should_cancel)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Ringkasan"
    used_titles = {"ringkasan"}
    summary_rows: list[list[Any]] = []
    page_count = 0
    table_count = 0
    data_row_count = 0
    empty_pages = 0
    used_generic = False  # True jika ada PDF non-Rincian (butuh sheet Ringkasan)

    for file_index, path in enumerate(paths, 1):
        ensure_not_cancelled(should_cancel)
        try:
            with pdfplumber.open(path) as pdf:
                total_pages = len(pdf.pages)
                if total_pages == 0:
                    summary_rows.append([path.name, "-", "Kosong", 0, "-"])
                    continue

                if _is_rincian_penjualan(pdf):
                    sheet_title, pages_written, rows_written = _write_rincian_penjualan_sheet(
                        workbook, pdf, path, used_titles, should_cancel
                    )
                    page_count += pages_written
                    table_count += 1
                    data_row_count += rows_written
                    summary_rows.append(
                        [path.name, pages_written, "Rincian Penjualan", rows_written, sheet_title]
                    )
                    continue

                used_generic = True
                for page_number, page in enumerate(pdf.pages, 1):
                    ensure_not_cancelled(should_cancel)
                    if progress:
                        progress(file_index, len(paths), path.name, page_number, total_pages)

                    page_count += 1
                    sheet_title = _unique_sheet_title(f"{path.stem} {page_number}", used_titles)
                    sheet = workbook.create_sheet(sheet_title)
                    header_rows: set[int] = set()
                    written_rows = 0
                    page_table_count = 0

                    extracted_tables = page.extract_tables() or []
                    for raw_table in extracted_tables:
                        table = [
                            [_safe_pdf_cell(cell) for cell in row]
                            for row in (raw_table or [])
                            if row and any(_safe_pdf_cell(cell) for cell in row)
                        ]
                        if not table:
                            continue
                        if written_rows:
                            sheet.append([])
                        next_row = 1 if not written_rows else sheet.max_row + 1
                        header_rows.add(next_row)
                        for row in table:
                            sheet.append(row)
                        written_rows += len(table)
                        data_row_count += len(table)
                        table_count += 1
                        page_table_count += 1

                    content_type = "Tabel"
                    if not written_rows:
                        text_rows = _pdf_text_rows(page.extract_text(layout=True) or "")
                        for row in text_rows:
                            sheet.append(row)
                        written_rows = len(text_rows)
                        data_row_count += written_rows
                        content_type = "Teks" if written_rows else "Tidak terbaca"

                    if not written_rows:
                        empty_pages += 1
                        sheet.append(["Tidak ada tabel atau teks yang dapat diekstrak dari halaman ini."])
                    _style_pdf_sheet(sheet, header_rows)
                    summary_rows.append(
                        [path.name, page_number, content_type, written_rows, sheet_title]
                    )
        except ProcessCancelled:
            raise
        except Exception as exc:
            raise ValueError(f"Gagal membaca {path.name}: {exc}") from exc

    if data_row_count == 0:
        raise ValueError(
            "Tidak ada teks atau tabel yang terbaca. PDF kemungkinan berupa hasil scan; gunakan PDF dengan teks digital atau lakukan OCR terlebih dahulu."
        )

    if used_generic:
        summary_headers = ["File PDF", "Halaman", "Isi Terdeteksi", "Jumlah Baris", "Sheet Excel"]
        summary.append(summary_headers)
        for row in summary_rows:
            summary.append(row)
        summary.sheet_view.showGridLines = False
        summary.freeze_panes = "A2"
        summary.auto_filter.ref = f"A1:E{summary.max_row}"
        for cell in summary[1]:
            cell.fill = PatternFill("solid", fgColor="117F74")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(vertical="center")
        for row in summary.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column, width in {"A": 38, "B": 12, "C": 18, "D": 15, "E": 31}.items():
            summary.column_dimensions[column].width = width
    elif len(workbook.sheetnames) > 1:
        # Semua PDF berupa Rincian Penjualan: buang sheet "Ringkasan" agar 1 sheet saja.
        workbook.remove(summary)

    output_path = output_path or unique_output_path(f"PDF to Excel {datetime.now():%Y-%m-%d %H%M}")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    ensure_not_cancelled(should_cancel)
    workbook.save(output_path)
    return output_path, {
        "filename": output_path.name,
        "file_count": len(paths),
        "page_count": page_count,
        "table_count": table_count,
        "row_count": data_row_count,
        "empty_pages": empty_pages,
    }


def build_workbook(
    source: dict[str, Any],
    report_date: date | None = None,
    report_month: date | None = None,
    report_dates: list[date] | None = None,
    include_cancelled: bool = False,
    include_failed: bool = False,
    should_cancel=None,
) -> tuple[Path, dict[str, Any]]:
    ensure_not_cancelled(should_cancel)
    raw_records = source["records"]
    records = filter_records(raw_records, report_date, report_month, report_dates)
    if not records:
        raise ValueError("Tidak ada data yang cocok dengan tanggal/bulan yang dipilih.")

    ensure_not_cancelled(should_cancel)
    included, walk_in, ayo, excluded = split_records(
        records,
        exclude_cancelled=not include_cancelled,
        exclude_failed=not include_failed,
    )
    included = sorted(included, key=record_sort_key)
    walk_in = sorted(walk_in, key=record_sort_key)
    ayo = sorted(ayo, key=record_sort_key)
    duplicate_booking_ids = find_duplicate_booking_ids(included)

    first_date = choose_report_month(records, report_date, report_month, report_dates)
    month_name, short_month, month_year = month_label(first_date)
    last_day = calendar.monthrange(first_date.year, first_date.month)[1]
    if report_dates:
        selected_dates = sorted(report_dates)
        first_selected = selected_dates[0]
        last_selected = selected_dates[-1]
        if first_selected == last_selected:
            period = f"Laporan Periode {month_name} {first_selected.day}, {first_selected.year}"
        else:
            period = (
                f"Laporan Periode {month_name} {first_selected.day}, {first_selected.year} - "
                f"{month_name} {last_selected.day}, {last_selected.year}"
            )
    else:
        period = f"Laporan Periode {month_name} 1, {first_date.year} - {month_name} {last_day}, {first_date.year}"
    monthly_output = report_date is None
    sheet_month = short_month if monthly_output else month_name

    wb = Workbook()
    wb.remove(wb.active)

    ensure_not_cancelled(should_cancel)
    write_summary_sheet(wb, f"Summary {short_month}", first_date, included, summary_dates=sorted(report_dates) if report_dates else None)

    raw_headers = source["headers"]
    walk_headers = [header for header in raw_headers if header != "No"]
    ensure_not_cancelled(should_cancel)
    write_table_sheet(
        wb,
        f"Walk In {sheet_month}",
        "OMSET SEWA LAPANGAN (Walk In)",
        f"Laporan Periode {month_year.upper()}",
        walk_headers,
        walk_in,
        "F59E0B",
        total_label_col=1,
        group_by_date=monthly_output,
    )

    ensure_not_cancelled(should_cancel)
    write_table_sheet(
        wb,
        f"AYO-{sheet_month}",
        "OMSET SEWA LAPANGAN (AYO)",
        period,
        raw_headers,
        ayo,
        "2563EB",
        total_label_col=1,
        group_by_date=monthly_output,
    )

    ensure_not_cancelled(should_cancel)
    write_all_sheet(wb, f"ALL-{sheet_month}", period, raw_headers, included, group_by_date=monthly_output)

    output_path = unique_output_path(build_omset_filename(report_date, report_month, report_dates))
    output_name = output_path.name
    ensure_not_cancelled(should_cancel)
    wb.save(output_path)

    ensure_not_cancelled(should_cancel)
    stats = {
        "filename": output_name,
        "month": month_year,
        "selected_date": report_date.isoformat() if report_date else "",
        "raw_rows": len(raw_records),
        "included_rows": len(included),
        "excluded_rows": len(excluded),
        "ayo_rows": len(ayo),
        "walk_in_rows": len(walk_in),
        "total_revenue": sum(number_value(r.get("Revenue Venue")) for r in included),
        "ayo_revenue": sum(number_value(r.get("Revenue Venue")) for r in ayo),
        "walk_in_revenue": sum(number_value(r.get("Revenue Venue")) for r in walk_in),
        "courts": sorted({normalize_header(r.get("Court")) for r in included if r.get("Court")}, key=court_sort_key),
        "duplicate_booking_ids": duplicate_booking_ids,
        "duplicate_booking_count": len(duplicate_booking_ids),
        "duplicate_booking_rows": sum(len(rows) for rows in duplicate_booking_ids.values()),
    }
    return output_path, stats


def write_category_summary_sheet(
    ws,
    file_results: list[dict[str, Any]],
    duplicates: dict[str, list[str]],
) -> None:
    """Tulis ringkasan omset keseluruhan per kategori ke worksheet ws."""
    duplicate_keys = {category.upper() for category in duplicates}
    headers = ["No", "File", "Kolom", "Group", "Tanggal", "Jumlah Data", "Nominal", "Status"]

    ws.title = "Omset Per Kategori"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "0F766E"

    style_title(
        ws,
        "OMSET KESELURUHAN PER KATEGORI",
        f"Dibuat: {datetime.now():%Y-%m-%d %H:%M}",
        len(headers),
    )

    header_row = 4
    for col, header in enumerate(headers, 1):
        ws.cell(header_row, col, header)
    style_header(ws, header_row, len(headers))

    row_idx = header_row + 1
    no = 1
    total_count = 0
    total_amount = 0.0
    for result in file_results:
        first_date = str(result.get("first_date") or "")
        last_date = str(result.get("last_date") or "")
        if first_date and last_date and first_date != last_date:
            date_label = f"{first_date} s/d {last_date}"
        else:
            date_label = first_date or last_date or "-"
        amounts = result.get("amounts") if isinstance(result.get("amounts"), dict) else {}
        for category, count in sorted(result["categories"].items()):
            amount = number_value(amounts.get(category, 0)) if isinstance(amounts, dict) else 0.0
            status = "Duplikat" if category.upper() in duplicate_keys else "OK"
            ws.cell(row_idx, 1, no)
            ws.cell(row_idx, 2, result["filename"])
            ws.cell(row_idx, 3, result.get("group_column") or "-")
            ws.cell(row_idx, 4, category)
            ws.cell(row_idx, 5, date_label)
            ws.cell(row_idx, 6, count)
            ws.cell(row_idx, 7, amount)
            ws.cell(row_idx, 8, status)
            total_count += count
            total_amount += amount
            no += 1
            row_idx += 1

    total_row = row_idx
    ws.cell(total_row, 4, "TOTAL")
    ws.cell(total_row, 6, total_count)
    ws.cell(total_row, 7, total_amount)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(total_row, col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E2E8F0")

    for row in range(header_row + 1, total_row + 1):
        ws.cell(row, 6).number_format = "#,##0"
        ws.cell(row, 7).number_format = "#,##0"

    widths = {1: 6, 2: 32, 3: 14, 4: 30, 5: 22, 6: 13, 7: 16, 8: 12}
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    apply_table_borders(ws, header_row, total_row, 1, len(headers))
    ws.freeze_panes = "A5"


def build_category_workbook(
    file_results: list[dict[str, Any]],
    duplicates: dict[str, list[str]],
    output_path: Path,
) -> Path:
    """Tulis rekap omset keseluruhan per kategori (1 sheet ringkasan) ke output_path."""
    wb = Workbook()
    write_category_summary_sheet(wb.active, file_results, duplicates)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


MONEY_DETAIL_KEYS = {"amount", "total", "total amount", "nominal", "price", "subtotal", "grand total"}


def read_category_detail(file_obj, filename: str, should_cancel=None) -> dict[str, Any]:
    """Baca seluruh baris data sebuah file kategori (header + isi) untuk jadi 1 sheet."""
    ensure_not_cancelled(should_cancel)
    wb = load_workbook(file_obj, data_only=True, read_only=True)
    ws = wb.active
    try:
        group_header, header_row, group_col_idx = find_any_column_position(ws, ("item group", "group", "kategori"))
    except ValueError as exc:
        wb.close()
        raise ValueError(f"{filename}: {exc}") from exc

    amount_column = find_optional_column_position(ws, ("amount", "total", "total amount"))
    date_column = find_optional_column_position(ws, ("order date", "tanggal", "date"))
    item_name_column = find_optional_column_position(ws, ITEM_NAME_COLUMN_ALIASES)

    header_values = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    raw_headers = [normalize_header(value) for value in header_values]
    last_col = max((idx for idx, header in enumerate(raw_headers) if header), default=0) + 1
    headers = raw_headers[:last_col]

    rows: list[list[Any]] = []
    for row_idx, row_values in enumerate(
        ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1
    ):
        if row_idx % 500 == 0:
            ensure_not_cancelled(should_cancel)
        cells = list(row_values[:last_col])
        if all(cell in (None, "") for cell in cells):
            continue
        rows.append([format_excel_value(cell) for cell in cells])

    wb.close()
    return {
        "filename": filename,
        "group_column": group_header,
        "headers": headers,
        "rows": rows,
        "group_col_idx": group_col_idx,
        "amount_col_idx": amount_column[2] if amount_column else None,
        "date_col_idx": date_column[2] if date_column else None,
        "item_name_col_idx": item_name_column[2] if item_name_column else None,
    }


def month_sheet_label(month: date) -> str:
    return f"{MONTH_ABBREV[month.month]}{month.year % 100:02d}"


def split_category_details_by_month(details: list[dict[str, Any]]) -> dict[date, list[dict[str, Any]]]:
    """Kelompokkan baris tiap file kategori per bulan (dibaca dari kolom tanggalnya)."""
    months: dict[date, dict[tuple[str, str], dict[str, Any]]] = {}
    for detail in details:
        group_idx = detail.get("group_col_idx")
        amount_idx = detail.get("amount_col_idx")
        date_idx = detail.get("date_col_idx")
        item_name_idx = detail.get("item_name_col_idx")
        parsed_rows: list[tuple[date | None, str, float, list[Any]]] = []
        file_months: set[date] = set()
        for row in detail["rows"]:
            group = normalize_header(row[group_idx]) if group_idx is not None and group_idx < len(row) else ""
            item_name = (
                row[item_name_idx]
                if item_name_idx is not None and item_name_idx < len(row)
                else None
            )
            category = category_from_group_and_item(group, item_name)
            if not category:
                continue  # baris total/footer bawaan file sumber
            row_date = parse_date(row[date_idx]) if date_idx is not None and date_idx < len(row) else None
            amount = (
                number_value(row[amount_idx])
                if amount_idx is not None and amount_idx < len(row)
                else 0.0
            )
            if row_date:
                file_months.add(date(row_date.year, row_date.month, 1))
            parsed_rows.append((row_date, category, amount, row))
        if not parsed_rows:
            continue
        if not file_months:
            raise ValueError(f"{detail['filename']}: kolom tanggal tidak terbaca, bulan tidak bisa ditentukan.")
        fallback_month = min(file_months)
        for row_date, category, amount, row in parsed_rows:
            month = date(row_date.year, row_date.month, 1) if row_date else fallback_month
            entry_key = (detail["filename"], category)
            entry = months.setdefault(month, {}).setdefault(
                entry_key,
                {
                    "filename": detail["filename"],
                    "category": category,
                    "headers": detail["headers"],
                    "group_header": detail.get("group_column") or "",
                    "rows": [],
                    "daily": defaultdict(float),
                    "total_amount": 0.0,
                },
            )
            entry["rows"].append(row)
            entry["total_amount"] += amount
            if row_date:
                entry["daily"][row_date] += amount
    return {
        month: sorted(by_file.values(), key=lambda e: str(e["category"]).upper())
        for month, by_file in months.items()
    }


def write_penjualan_per_kategori_sheet(
    wb: Workbook,
    sheet_name: str,
    month: date,
    entries: list[dict[str, Any]],
) -> None:
    """Sheet 'Penjualan per Kategori': judul bulan + header Olsera, baris dikelompokkan per
    kategori dengan nama kategori di kolom A (di-merge). Mengikuti format file contoh:
    tanpa subtotal/total, semua kolom sumber dipertahankan."""
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "0F766E"

    # Master header = gabungan semua kolom file (asumsi seragam), semua kolom dipertahankan.
    headers = ["KATEGORI"]
    for entry in entries:
        for header in entry["headers"]:
            if header and header not in headers:
                headers.append(header)

    month_title = f"{MONTH_NAMES[month.month].upper()}'{month.year % 100:02d}"
    title_cell = ws.cell(1, 1, month_title)
    title_cell.font = Font(bold=True, size=12)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(3, len(headers)))

    header_row = 2
    for col, header in enumerate(headers, 1):
        ws.cell(header_row, col, header)
    style_header(ws, header_row, len(headers))
    ws.freeze_panes = "A3"

    money_cols = {col for col, header in enumerate(headers, 1) if as_key(header) in MONEY_DETAIL_KEYS}
    row_idx = header_row + 1
    for entry in entries:
        index_by_header = {header: idx for idx, header in enumerate(entry["headers"]) if header}
        block_start = row_idx
        for offset, row in enumerate(entry["rows"]):
            if offset == 0:
                ws.cell(row_idx, 1, entry["category"])
            for col, header in enumerate(headers[1:], 2):
                idx = index_by_header.get(header)
                if idx is not None and idx < len(row):
                    ws.cell(row_idx, col, row[idx])
            row_idx += 1
        block_end = row_idx - 1
        if block_end > block_start:
            ws.merge_cells(start_row=block_start, start_column=1, end_row=block_end, end_column=1)
        ws.cell(block_start, 1).alignment = Alignment(vertical="center")

    last_row = row_idx - 1
    for row in range(header_row + 1, last_row + 1):
        for col in money_cols:
            ws.cell(row, col).number_format = "#,##0"

    ws.column_dimensions["A"].width = 24
    for col in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16

    apply_table_borders(ws, header_row, max(header_row, last_row), 1, len(headers))


def write_omset_sheet(wb: Workbook, sheet_name: str, month: date, entries: list[dict[str, Any]]) -> None:
    """Sheet 'OMSET': matriks kategori (baris) x hari 1..akhir-bulan (kolom), nilai = omset
    harian per kategori, kolom TOTAL + baris TOTAL. Format mengikuti file contoh."""
    last_day = calendar.monthrange(month.year, month.month)[1]

    daily_by_category: dict[str, dict[date, float]] = {}
    for entry in entries:
        bucket = daily_by_category.setdefault(str(entry["category"]), defaultdict(float))
        for day, amount in entry["daily"].items():
            bucket[day] += amount
    categories = sorted(daily_by_category, key=str.upper)

    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "F59E0B"

    first_day_col = 3                      # kolom C = hari ke-1
    last_day_col = first_day_col + last_day - 1
    total_col = last_day_col + 1

    title_cell = ws.cell(1, 1, "OMSET KESELURUHAN")
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="174E57")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_col)
    subtitle = ws.cell(2, 1, f"{MONTH_NAMES[month.month].upper()}'{month.year % 100:02d}")
    subtitle.font = Font(bold=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_col)

    hr1, hr2 = 4, 5
    ws.cell(hr1, 1, "No")
    ws.merge_cells(start_row=hr1, start_column=1, end_row=hr2, end_column=1)
    ws.cell(hr1, 2, "Kategori")
    ws.merge_cells(start_row=hr1, start_column=2, end_row=hr2, end_column=2)
    ws.cell(hr1, first_day_col, "Tanggal")
    ws.merge_cells(start_row=hr1, start_column=first_day_col, end_row=hr1, end_column=last_day_col)
    for offset in range(last_day):
        ws.cell(hr2, first_day_col + offset, offset + 1)
    ws.cell(hr1, total_col, "TOTAL")
    ws.merge_cells(start_row=hr1, start_column=total_col, end_row=hr2, end_column=total_col)
    style_header(ws, hr1, total_col)
    style_header(ws, hr2, total_col)
    ws.freeze_panes = "C6"

    first_letter = get_column_letter(first_day_col)
    last_letter = get_column_letter(last_day_col)
    row_idx = hr2 + 1
    for no, category in enumerate(categories, 1):
        ws.cell(row_idx, 1, no)
        ws.cell(row_idx, 2, category)
        bucket = daily_by_category[category]
        for offset in range(last_day):
            amount = bucket.get(date(month.year, month.month, offset + 1), 0)
            if amount:
                ws.cell(row_idx, first_day_col + offset, amount)
        ws.cell(row_idx, total_col, f"=SUM({first_letter}{row_idx}:{last_letter}{row_idx})")
        row_idx += 1

    total_row = row_idx
    ws.cell(total_row, 2, "TOTAL")
    if categories:
        for col in range(first_day_col, total_col + 1):
            letter = get_column_letter(col)
            ws.cell(total_row, col, f"=SUM({letter}{hr2 + 1}:{letter}{total_row - 1})")
    for col in range(1, total_col + 1):
        cell = ws.cell(total_row, col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E2E8F0")

    for row in range(hr2 + 1, total_row + 1):
        for col in range(first_day_col, total_col + 1):
            ws.cell(row, col).number_format = "#,##0"

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 26
    for col in range(first_day_col, total_col):
        ws.column_dimensions[get_column_letter(col)].width = 10
    ws.column_dimensions[get_column_letter(total_col)].width = 14
    apply_table_borders(ws, hr1, total_row, 1, total_col)


OLSERA_COURT_TOKENS = ("LAPANGAN", "COURT")


def olsera_court_daily(entries: list[dict[str, Any]]) -> tuple[dict[date, float], dict[date, float]]:
    """Omset harian kategori sewa lapangan dari data Olsera, dipisah (padel, pickle)."""
    padel: dict[date, float] = defaultdict(float)
    pickle: dict[date, float] = defaultdict(float)
    for entry in entries:
        name = str(entry["category"]).upper()
        if not any(token in name for token in OLSERA_COURT_TOKENS):
            continue
        target = pickle if "PICKLE" in name else padel
        for day, amount in entry["daily"].items():
            target[day] += amount
    return padel, pickle


def write_summary_lapangan_sheet(
    wb: Workbook,
    sheet_name: str,
    month: date,
    records: list[dict[str, Any]],
    olsera_padel: dict[date, float],
    olsera_pickle: dict[date, float],
) -> None:
    """Sheet 'Summary Lapangan': summary booking per court + kolom AYO/OLSERA/SELISIH."""
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "174E57"

    ws["A1"] = "SUMMARY LAPANGAN PADEL & PICKLE"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="174E57")
    ws["A2"] = f"Bulan: {MONTH_ABBREV[month.month].upper()} {month.year}"
    ws["A2"].font = Font(bold=True)

    summary: dict[tuple[date, str, str], float] = defaultdict(float)
    booking_dates: set[date] = set()
    for record in records:
        booking_date = parse_date(record.get("Date of Booking"))
        court = normalize_header(record.get("Court"))
        if not booking_date or not court:
            continue
        booking_dates.add(booking_date)
        summary[(booking_date, court, booking_channel(record))] += number_value(record.get("Revenue Venue"))

    courts = sorted({court for (_, court, _) in summary}, key=court_sort_key)
    olsera_daily: dict[date, float] = defaultdict(float)
    for source in (olsera_padel, olsera_pickle):
        for day, amount in source.items():
            olsera_daily[day] += amount

    header_row_1, header_row_2 = 4, 5
    total_col = 2 + 2 * len(courts)
    ayo_col, olsera_col, selisih_col = total_col + 1, total_col + 2, total_col + 3

    ws.cell(header_row_1, 1, "Date of\nBooking")
    ws.merge_cells(start_row=header_row_1, start_column=1, end_row=header_row_2, end_column=1)
    col = 2
    for court in courts:
        ws.cell(header_row_1, col, display_court_label(court))
        ws.merge_cells(start_row=header_row_1, start_column=col, end_row=header_row_1, end_column=col + 1)
        ws.cell(header_row_2, col, "AYO")
        ws.cell(header_row_2, col + 1, "Walk In")
        col += 2
    for label, col_idx in (("TOTAL", total_col), ("AYO", ayo_col), ("OLSERA", olsera_col), ("SELISIH", selisih_col)):
        ws.cell(header_row_1, col_idx, label)
        ws.merge_cells(start_row=header_row_1, start_column=col_idx, end_row=header_row_2, end_column=col_idx)
    style_header(ws, header_row_1, selisih_col)
    style_header(ws, header_row_2, selisih_col)
    ws.freeze_panes = "B6"

    all_dates = booking_dates | set(olsera_daily)
    if all_dates:
        first_day, last_day = min(all_dates), max(all_dates)
        days = [first_day + timedelta(days=offset) for offset in range((last_day - first_day).days + 1)]
    else:
        days = []

    start_row = 6
    for row_offset, current_day in enumerate(days):
        row_idx = start_row + row_offset
        ws.cell(row_idx, 1, current_day)
        walk_in_total = 0.0
        col = 2
        for court in courts:
            ayo_amount = summary[(current_day, court, "AYO")]
            walk_amount = summary[(current_day, court, "Walk In")]
            ws.cell(row_idx, col, ayo_amount or None)
            ws.cell(row_idx, col + 1, walk_amount or None)
            walk_in_total += walk_amount
            col += 2
        if courts:
            ws.cell(row_idx, total_col, f"=SUM(B{row_idx}:{get_column_letter(total_col - 1)}{row_idx})")
        olsera_amount = olsera_daily.get(current_day, 0)
        ws.cell(row_idx, ayo_col, walk_in_total or None)
        ws.cell(row_idx, olsera_col, olsera_amount or None)
        if walk_in_total or olsera_amount:
            ws.cell(row_idx, selisih_col, olsera_amount - walk_in_total)

    total_row = start_row + len(days)
    ws.cell(total_row, 1, "TOTAL")
    if days:
        for col_idx in range(2, selisih_col + 1):
            letter = get_column_letter(col_idx)
            ws.cell(total_row, col_idx, f"=SUM({letter}{start_row}:{letter}{total_row - 1})")
    for col_idx in range(1, selisih_col + 1):
        cell = ws.cell(total_row, col_idx)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E2E8F0")

    padel_ayo = padel_walk = pickle_ayo = pickle_walk = 0.0
    for (_, court, channel), amount in summary.items():
        is_pickle = "PICKLE" in court.upper()
        if channel == "AYO":
            if is_pickle:
                pickle_ayo += amount
            else:
                padel_ayo += amount
        else:
            if is_pickle:
                pickle_walk += amount
            else:
                padel_walk += amount
    olsera_padel_total = sum(olsera_padel.values())
    olsera_pickle_total = sum(olsera_pickle.values())

    notes_start = total_row + 2
    label_col, padel_col, pickle_col = 2, 3, 4
    for col_idx, label in ((padel_col, "PADEL"), (pickle_col, "PICKLE")):
        cell = ws.cell(notes_start, col_idx, label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="174E57")
        cell.alignment = Alignment(horizontal="center")
    ws.cell(notes_start + 1, label_col, "AYO").font = Font(bold=True)
    ws.cell(notes_start + 1, padel_col, padel_ayo)
    ws.cell(notes_start + 1, pickle_col, pickle_ayo)
    ws.cell(notes_start + 2, label_col, "Walk In").font = Font(bold=True)
    ws.cell(notes_start + 2, padel_col, padel_walk)
    ws.cell(notes_start + 2, pickle_col, pickle_walk)

    compare_start = notes_start + 4
    for col_idx, label in ((padel_col, "Booking (AYO)"), (pickle_col, "Olsera"), (pickle_col + 1, "Selisih")):
        cell = ws.cell(compare_start, col_idx, label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="174E57")
        cell.alignment = Alignment(horizontal="center")
    ws.cell(compare_start + 1, label_col, "Walk In Padel").font = Font(bold=True)
    ws.cell(compare_start + 1, padel_col, padel_walk)
    ws.cell(compare_start + 1, pickle_col, olsera_padel_total)
    ws.cell(compare_start + 1, pickle_col + 1, olsera_padel_total - padel_walk)
    ws.cell(compare_start + 2, label_col, "Walk In Pickle").font = Font(bold=True)
    ws.cell(compare_start + 2, padel_col, pickle_walk)
    ws.cell(compare_start + 2, pickle_col, olsera_pickle_total)
    ws.cell(compare_start + 2, pickle_col + 1, olsera_pickle_total - pickle_walk)

    box_label_col = max(6, selisih_col - 1)
    ws.cell(notes_start + 1, box_label_col, "AYO").font = Font(bold=True)
    ws.cell(notes_start + 1, box_label_col + 1, padel_ayo + pickle_ayo)
    ws.cell(notes_start + 2, box_label_col, "Walk In").font = Font(bold=True)
    ws.cell(notes_start + 2, box_label_col + 1, padel_walk + pickle_walk)

    for row_offset in range(len(days)):
        ws.cell(start_row + row_offset, 1).number_format = "dd-mmm-yy"
    for row in range(start_row, total_row + 1):
        for col_idx in range(2, selisih_col + 1):
            ws.cell(row, col_idx).number_format = "#,##0"
    for row in range(notes_start + 1, compare_start + 3):
        for col_idx in (padel_col, pickle_col, pickle_col + 1, box_label_col + 1):
            ws.cell(row, col_idx).number_format = "#,##0"

    ws.column_dimensions["A"].width = 18
    for col_idx in range(2, max(selisih_col, box_label_col + 1) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15


def build_olah_data_workbook(details: list[dict[str, Any]], output_path: Path) -> Path:
    """Olah semua file kategori jadi 1 workbook: 2 sheet per bulan.

    Per bulan: 'Penjualan per Kategori - <Bln><Thn>' dan 'OMSET - <Bln><Thn>'.
    (Sheet 'Summary Lapangan' sengaja tidak dibuat di fitur olah data per kategori.)
    """
    months = split_category_details_by_month(details)
    if not months:
        raise ValueError("Tidak ada data kategori yang bisa diolah.")

    wb = Workbook()
    wb.remove(wb.active)
    for month in sorted(months):
        label = month_sheet_label(month)
        entries = months[month]
        write_penjualan_per_kategori_sheet(wb, f"Penjualan per Kategori - {label}", month, entries)
        write_omset_sheet(wb, f"OMSET - {label}", month, entries)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
